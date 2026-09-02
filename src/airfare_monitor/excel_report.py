"""Generate the consolidated Excel workbook."""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import LegResult, RunReport

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_HIT_FILL = PatternFill("solid", fgColor="C6EFCE")
_FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
_CURRENCY_FORMAT = '¥#,##0.00;[Red]-¥#,##0.00'


def _money(value: Decimal | None) -> float | None:
    return float(value) if value is not None else None


def _delta(current: Decimal | None, reference: Decimal | None) -> Decimal | None:
    return current - reference if current is not None and reference is not None else None


def _style_sheet(sheet: Any, *, freeze: str = "A2") -> None:
    sheet.freeze_panes = freeze
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column in sheet.columns:
        values = [str(cell.value or "") for cell in column[:100]]
        width = min(max(max(map(len, values), default=8) + 2, 10), 28)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = width


def _summary_sheet(workbook: Workbook, report: RunReport) -> None:
    sheet = workbook.active
    sheet.title = "本次汇总"
    headers = [
        "航程",
        "行程类型",
        "出发地",
        "到达地",
        "去程日期",
        "去程ETD时间窗",
        "返程日期",
        "返程ETD时间窗",
        "心理价位",
        "本次最低总价",
        "与阈值差额",
        "较上次变化",
        "是否命中",
        "符合航班数",
        "采集状态",
        "采集时间",
        "错误说明",
    ]
    sheet.append(headers)
    for result in report.legs:
        minimum = result.minimum_total_cny
        row = [
            result.leg.id,
            "往返" if result.leg.is_round_trip else "单程",
            result.leg.origin_display,
            result.leg.destination_display,
            result.leg.departure_date,
            result.leg.etd_window.display(),
            result.leg.return_date,
            result.leg.return_etd_window.display() if result.leg.return_etd_window else None,
            _money(result.leg.expected_total_price_cny),
            _money(minimum),
            _money(_delta(minimum, result.leg.expected_total_price_cny)),
            _money(_delta(minimum, result.previous_min_total_cny)),
            "是（已确认）" if result.leg.id in report.threshold_confirmed_leg_ids else "否",
            result.eligible_count,
            str(result.status),
            result.captured_at,
            result.error_message,
        ]
        sheet.append(row)
        row_number = sheet.max_row
        for column in (9, 10, 11, 12):
            sheet.cell(row_number, column).number_format = _CURRENCY_FORMAT
        sheet.cell(row_number, 5).number_format = "yyyy-mm-dd"
        sheet.cell(row_number, 7).number_format = "yyyy-mm-dd"
        sheet.cell(row_number, 16).number_format = "yyyy-mm-dd hh:mm"
        fill = _HIT_FILL if result.leg.id in report.threshold_confirmed_leg_ids else None
        if str(result.status) != "success":
            fill = _FAIL_FILL
        if fill:
            for cell in sheet[row_number]:
                cell.fill = fill
    _style_sheet(sheet)


def _leg_sheet(workbook: Workbook, result: LegResult, index: int) -> None:
    sheet = workbook.create_sheet(f"航程{index}")
    headers = [
        "排名",
        "航班号",
        "航司",
        "出发机场",
        "到达机场",
        "出发日期",
        "ETD",
        "ETA",
        "飞行时长(分钟)",
        "航段数",
        "行程类型",
        "中转机场",
        "中转等待(分钟)",
        "基础票价",
        "税费",
        "总价",
        "币种",
        "余票提示",
        "免费行李件数",
        "免费行李重量",
        "报价来源",
        "采集时间",
        "返程航班号",
        "返程航司",
        "返程出发机场",
        "返程到达机场",
        "返程日期",
        "返程ETD",
        "返程ETA",
        "返程飞行时长(分钟)",
        "返程航段数",
        "返程类型",
        "返程中转机场",
        "返程中转等待(分钟)",
        "去程余票提示",
        "返程余票提示",
        "航班签名",
    ]
    sheet.append(headers)
    for rank, flight in enumerate(result.flights, start=1):
        inbound = flight.return_itinerary
        sheet.append(
            [
                rank,
                flight.flight_codes_display,
                flight.carrier_codes_display,
                flight.origin_airport_iata,
                flight.destination_airport_iata,
                flight.departure_date,
                flight.etd_local,
                flight.eta_local,
                flight.duration_minutes,
                flight.segment_count,
                "直达" if flight.is_direct else f"中转{flight.segment_count - 1}次",
                flight.connection_airports_display,
                flight.layover_minutes,
                _money(flight.base_price_cny),
                _money(flight.tax_cny),
                _money(flight.total_price_cny),
                flight.currency_code,
                (
                    flight.seat_availability.display()
                    if flight.seat_availability
                    else flight.remaining_seats
                ),
                flight.free_baggage_piece,
                flight.free_baggage_weight,
                flight.source_domain,
                flight.captured_at,
                inbound.flight_codes_display if inbound else None,
                inbound.carrier_codes_display if inbound else None,
                inbound.origin_airport_iata if inbound else None,
                inbound.destination_airport_iata if inbound else None,
                inbound.departure_date if inbound else None,
                inbound.etd_local if inbound else None,
                inbound.eta_local if inbound else None,
                inbound.duration_minutes if inbound else None,
                inbound.segment_count if inbound else None,
                ("直达" if inbound.is_direct else f"中转{inbound.segment_count - 1}次") if inbound else None,
                inbound.connection_airports_display if inbound else None,
                inbound.layover_minutes if inbound else None,
                (
                    flight.outbound_seat_availability.display()
                    if flight.outbound_seat_availability
                    else flight.remaining_seats
                ),
                (
                    inbound.seat_availability.display()
                    if inbound and inbound.seat_availability
                    else None
                ),
                flight.flight_signature,
            ]
        )
        for column in (14, 15, 16):
            sheet.cell(sheet.max_row, column).number_format = _CURRENCY_FORMAT
        for column in (6, 7, 8, 22, 27, 28, 29):
            sheet.cell(sheet.max_row, column).number_format = "yyyy-mm-dd hh:mm"
    _style_sheet(sheet)
    sheet.column_dimensions[get_column_letter(len(headers))].hidden = True


def _history_sheet(workbook: Workbook, history: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("24小时历史")
    sheet.append(["航程", "类型", "航线", "采集时间", "最低总价", "采集状态"])
    for item in history:
        value = item.get("minimum_total_price_cny")
        is_round_trip = bool(item.get("return_date"))
        arrow = "↔" if is_round_trip else "-"
        sheet.append(
            [
                item["leg_id"],
                "往返" if is_round_trip else "单程",
                f'{item["origin_airport_iata"]}{arrow}{item["destination_airport_iata"]}',
                datetime.fromisoformat(item["captured_at"]),
                float(value) if value is not None else None,
                item["status"],
            ]
        )
        sheet.cell(sheet.max_row, 4).number_format = "yyyy-mm-dd hh:mm"
        sheet.cell(sheet.max_row, 5).number_format = _CURRENCY_FORMAT
    _style_sheet(sheet)

def generate_workbook(report: RunReport, history: list[dict[str, Any]], output_directory: str | Path) -> Path:
    output = Path(output_directory)
    output.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    _summary_sheet(workbook, report)
    for index, result in enumerate(report.legs, start=1):
        _leg_sheet(workbook, result, index)
    _history_sheet(workbook, history)
    filename = f"airfare-monitor_{report.finished_at:%Y%m%d_%H%M}.xlsx"
    destination = output / filename
    workbook.save(destination)
    return destination
