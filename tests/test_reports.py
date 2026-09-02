from __future__ import annotations

import json
import sqlite3
import tempfile
import unittest
from contextlib import closing
from dataclasses import replace
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from airfare_monitor.config import MailSettings
from airfare_monitor.excel_report import generate_workbook
from airfare_monitor.mail import build_message, build_subject
from airfare_monitor.models import (
    EtdWindow,
    FlightSnapshot,
    ItinerarySnapshot,
    LegConfig,
    LegResult,
    LegStatus,
    PreferredPriceReference,
    PreferredSchedule,
    RunReport,
    RunStatus,
    SeatAvailability,
)
from airfare_monitor.storage import SQLiteStore


def route() -> LegConfig:
    return LegConfig("leg-1", True, "SHA", "KUL", date(2026, 9, 27), EtdWindow(time(0), time(23, 59)), True, Decimal("1000"), 10, 1, 0, "economy", "上海", "吉隆坡")


def report(status: RunStatus = RunStatus.SUCCESS) -> RunReport:
    captured = datetime(2026, 8, 31, 9, 30)
    result = LegResult(route(), LegStatus.SUCCESS, captured, completed_response=True)
    return RunReport("run-1", captured, captured, status, [result], set())


MAIL = MailSettings(False, "smtp.example.com", 465, "ssl", "U", "P", "S", "R", True, "[航价监控]", "[低价命中]")


class ReportTests(unittest.TestCase):
    def test_subject_uses_enabled_leg_count(self):
        self.assertIn("1程更新", build_subject(report(), MAIL))

    def test_subject_marks_partial(self):
        self.assertIn("[部分失败]", build_subject(report(RunStatus.PARTIAL), MAIL))

    def test_message_has_plain_html_and_attachment(self):
        with tempfile.TemporaryDirectory() as directory:
            attachment = Path(directory) / "report.xlsx"
            attachment.write_bytes(b"test")
            message = build_message(report(), MAIL, sender="sender@example.com", recipients=["to@example.com"], attachment=attachment)
            self.assertTrue(message.is_multipart())
            self.assertEqual(len(list(message.iter_attachments())), 1)
            self.assertIn("上海（SHA） → 吉隆坡（KUL）", message.get_body(preferencelist=("plain",)).get_content())

    def test_message_includes_preferred_schedule_live_price(self):
        preference = PreferredSchedule(
            "上海浦东 → 吉隆坡",
            time(7, 25),
            time(13),
            departure_tolerance_minutes=60,
            arrival_tolerance_minutes=60,
            origin_airport_iata="PVG",
            destination_airport_iata="KUL",
        )
        configured = replace(route(), preferred_schedules=(preference,))
        captured = datetime(2026, 8, 31, 9, 30)
        flight = FlightSnapshot(
            flight_signature="preferred-1",
            flight_codes=("MU001",),
            carrier_codes=("MU",),
            origin_airport_iata="PVG",
            destination_airport_iata="KUL",
            departure_date=date(2026, 9, 27),
            etd_local=datetime(2026, 9, 27, 7, 40),
            eta_local=datetime(2026, 9, 27, 13, 15),
            duration_minutes=335,
            segment_count=1,
            is_direct=True,
            base_price_cny=Decimal("800"),
            tax_cny=Decimal("260"),
            total_price_cny=Decimal("1060"),
            currency_code="CNY",
            remaining_seats=None,
            free_baggage_piece=None,
            free_baggage_weight=None,
            source_domain=None,
            captured_at=captured,
        )
        result = LegResult(
            configured,
            LegStatus.SUCCESS,
            captured,
            flights=[flight],
            preferred_matches=[flight],
            preferred_price_references=[
                PreferredPriceReference(
                    first_total_price_cny=Decimal("1200"),
                    first_captured_at=datetime(2026, 8, 30, 9, 30),
                    previous_total_price_cny=Decimal("1100"),
                    previous_captured_at=datetime(2026, 8, 31, 9),
                )
            ],
            completed_response=True,
        )
        value = RunReport("run-1", captured, captured, RunStatus.SUCCESS, [result], set())
        message = build_message(value, MAIL, sender="sender@example.com", recipients=["to@example.com"])
        body = message.get_body(preferencelist=("plain",)).get_content()
        self.assertIn("重点关注时段价格", body)
        self.assertIn("上海浦东 → 吉隆坡", body)
        self.assertIn("实际 09-27 07:40 → 09-27 13:15", body)
        self.assertIn("本次：¥1,060", body)
        self.assertIn("首次：¥1,200", body)
        self.assertIn("较首次：下降 ¥140", body)
        self.assertIn("上次：¥1,100", body)
        self.assertIn("较上次：下降 ¥40", body)
        self.assertLess(body.index("重点关注时段价格"), body.index("最低价候选"))

        with tempfile.TemporaryDirectory() as directory:
            store = SQLiteStore(Path(directory) / "monitor.sqlite3")
            store.initialize()
            store.save_report(value)
            first_reference = store.preferred_price_reference(
                "leg-1", preference.history_key(configured.departure_date), before=datetime(2026, 8, 31, 10)
            )
            self.assertEqual(first_reference.first_total_price_cny, Decimal("1060"))
            self.assertEqual(first_reference.previous_total_price_cny, Decimal("1060"))

            later = datetime(2026, 8, 31, 10, 30)
            cheaper = replace(flight, total_price_cny=Decimal("990"), captured_at=later)
            later_result = LegResult(
                configured,
                LegStatus.SUCCESS,
                later,
                flights=[cheaper],
                preferred_matches=[cheaper],
                completed_response=True,
            )
            store.save_report(RunReport("run-2", later, later, RunStatus.SUCCESS, [later_result], set()))
            later_reference = store.preferred_price_reference(
                "leg-1", preference.history_key(configured.departure_date), before=datetime(2026, 8, 31, 11)
            )
            self.assertEqual(later_reference.first_total_price_cny, Decimal("1060"))
            self.assertEqual(later_reference.previous_total_price_cny, Decimal("990"))

    def test_domestic_mail_shows_fare_fees_and_estimated_payment(self):
        captured = datetime(2026, 9, 1, 11, 8)
        domestic = replace(
            route(),
            id="leg-5",
            destination_airport_iata="XMN",
            destination_name_zh="厦门",
            departure_date=date(2026, 9, 29),
            market="auto",
        )
        flight = FlightSnapshot(
            flight_signature="domestic-1",
            flight_codes=("9C8815",),
            carrier_codes=("9C",),
            origin_airport_iata="SHA",
            destination_airport_iata="XMN",
            departure_date=date(2026, 9, 29),
            etd_local=datetime(2026, 9, 29, 7, 15),
            eta_local=datetime(2026, 9, 29, 9, 5),
            duration_minutes=110,
            segment_count=1,
            is_direct=True,
            base_price_cny=Decimal("350"),
            tax_cny=Decimal("120"),
            total_price_cny=Decimal("470"),
            currency_code="CNY",
            remaining_seats=None,
            free_baggage_piece=None,
            free_baggage_weight=None,
            source_domain="ly.com",
            captured_at=captured,
        )
        result = LegResult(domestic, LegStatus.SUCCESS, captured, flights=[flight], completed_response=True)
        value = RunReport("run-domestic", captured, captured, RunStatus.SUCCESS, [result], set())
        message = build_message(value, MAIL, sender="sender@example.com", recipients=["to@example.com"])
        expected = "票面 ¥350 + 机建燃油 ¥120 = 预计支付 ¥470"
        plain = message.get_body(preferencelist=("plain",)).get_content()
        html_body = message.get_body(preferencelist=("html",)).get_content()
        self.assertGreaterEqual(plain.count(expected), 2)
        self.assertIn(expected, html_body)
        self.assertIn("国内航班同时显示票面价、机建燃油和预计支付总价", plain)

    def test_transfer_mail_shows_connection_and_observation_only_threshold(self):
        captured = datetime(2026, 9, 2, 9, 30)
        configured = replace(
            route(),
            direct_only=False,
            expected_total_price_cny=None,
            max_layover_minutes=240,
        )
        flight = FlightSnapshot(
            flight_signature="transfer-1",
            flight_codes=("MH001", "UL002"),
            carrier_codes=("MH", "UL"),
            origin_airport_iata="KUL",
            destination_airport_iata="MLE",
            departure_date=date(2026, 9, 27),
            etd_local=datetime(2026, 9, 27, 8),
            eta_local=datetime(2026, 9, 27, 16),
            duration_minutes=480,
            segment_count=2,
            is_direct=False,
            base_price_cny=Decimal("800"),
            tax_cny=Decimal("200"),
            total_price_cny=Decimal("1000"),
            currency_code="CNY",
            remaining_seats=None,
            free_baggage_piece=None,
            free_baggage_weight=None,
            source_domain="example.test",
            captured_at=captured,
            connection_airports=("CMB",),
            layover_minutes=120,
        )
        result = LegResult(configured, LegStatus.SUCCESS, captured, flights=[flight], completed_response=True)
        value = RunReport("run-transfer", captured, captured, RunStatus.SUCCESS, [result], set())
        self.assertFalse(result.threshold_hit)
        message = build_message(value, MAIL, sender="sender@example.com", recipients=["to@example.com"])
        plain = message.get_body(preferencelist=("plain",)).get_content()
        html_body = message.get_body(preferencelist=("html",)).get_content()
        self.assertIn("心理价位：—", plain)
        self.assertIn("中转1次，经 CMB，等待 2小时", plain)
        self.assertIn("中转1次，经 CMB，等待 2小时", html_body)

        with tempfile.TemporaryDirectory() as directory:
            store = SQLiteStore(Path(directory) / "monitor.sqlite3")
            store.initialize()
            store.save_report(value)
            with closing(store.connect()) as connection:
                row = connection.execute(
                    "SELECT connection_airports_json, layover_minutes FROM flight_snapshots"
                ).fetchone()
            self.assertEqual(row["connection_airports_json"], '["CMB"]')
            self.assertEqual(row["layover_minutes"], 120)
            output = generate_workbook(value, store.history(since=datetime(2026, 9, 1)), Path(directory))
            workbook = load_workbook(output, read_only=True)
            try:
                values = list(workbook["航程1"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
                self.assertEqual(values[10:13], ("中转1次", "CMB", 120))
            finally:
                workbook.close()

    def test_round_trip_mail_storage_and_workbook_show_combination_total(self):
        captured = datetime(2026, 9, 2, 10)
        configured = replace(
            route(),
            id="roundtrip-sha-kul",
            expected_total_price_cny=None,
            return_date=date(2026, 10, 3),
            return_etd_window=EtdWindow(time(0), time(23, 59)),
            return_direct_only=True,
        )
        inbound = ItinerarySnapshot(
            "inbound-signature",
            ("MU002",),
            ("MU",),
            "KUL",
            "PVG",
            date(2026, 10, 3),
            datetime(2026, 10, 3, 14),
            datetime(2026, 10, 3, 19, 20),
            320,
            1,
            True,
            seat_availability=SeatAvailability(2, "2张", "票少", False),
        )
        flight = FlightSnapshot(
            "roundtrip-signature",
            ("MU001",),
            ("MU",),
            "PVG",
            "KUL",
            date(2026, 9, 27),
            datetime(2026, 9, 27, 8),
            datetime(2026, 9, 27, 13, 30),
            330,
            1,
            True,
            Decimal("1200"),
            Decimal("600"),
            Decimal("1800"),
            "CNY",
            None,
            None,
            None,
            None,
            captured,
            return_itinerary=inbound,
            seat_availability=SeatAvailability(2, "2张", "票少", False),
            outbound_seat_availability=SeatAvailability(9, None, None, False),
        )
        result = LegResult(configured, LegStatus.SUCCESS, captured, flights=[flight], completed_response=True)
        value = RunReport("run-roundtrip", captured, captured, RunStatus.SUCCESS, [result], set())
        message = build_message(value, MAIL, sender="sender@example.com", recipients=["to@example.com"])
        plain = message.get_body(preferencelist=("plain",)).get_content()
        self.assertIn("1组往返更新", message["Subject"])
        self.assertIn("上海（SHA） ↔ 吉隆坡（KUL）", plain)
        self.assertIn("去程 MU001", plain)
        self.assertIn("返程 MU002", plain)
        self.assertIn("往返合计 ¥1,800", plain)
        expected_seats = "余票：整体 2张（票少） · 去程 9张或以上（平台提示） · 返程 2张（票少）"
        self.assertIn(expected_seats, plain)
        self.assertIn(expected_seats, message.get_body(preferencelist=("html",)).get_content())

        with tempfile.TemporaryDirectory() as directory:
            store = SQLiteStore(Path(directory) / "monitor.sqlite3")
            store.initialize()
            store.save_report(value)
            with closing(store.connect()) as connection:
                row = connection.execute(
                    """SELECT return_itinerary_json, seat_availability_json,
                                      outbound_seat_availability_json
                               FROM flight_snapshots"""
                ).fetchone()
            stored = json.loads(row["return_itinerary_json"])
            self.assertEqual(stored["flight_codes"], ["MU002"])
            self.assertEqual(stored["seat_availability"]["count_hint"], 2)
            self.assertEqual(json.loads(row["seat_availability_json"])["scarcity_text"], "票少")
            self.assertEqual(json.loads(row["outbound_seat_availability_json"])["count_hint"], 9)
            output = generate_workbook(value, [], Path(directory))
            workbook = load_workbook(output, read_only=True)
            try:
                summary = list(workbook["本次汇总"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
                self.assertEqual(summary[1], "往返")
                self.assertEqual(summary[6].date(), date(2026, 10, 3))
                values = list(workbook["航程1"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
                self.assertEqual(values[22], "MU002")
                self.assertEqual(values[24:26], ("KUL", "PVG"))
                self.assertEqual(values[34], "9张或以上（平台提示）")
                self.assertEqual(values[35], "2张（票少）")
            finally:
                workbook.close()

    def test_mail_shows_five_candidates_with_price_history_without_preferences(self):
        captured = datetime(2026, 9, 2, 11)
        base = FlightSnapshot(
            "candidate-1",
            ("MU001",),
            ("MU",),
            "PVG",
            "KUL",
            date(2026, 9, 27),
            datetime(2026, 9, 27, 8),
            datetime(2026, 9, 27, 13, 30),
            330,
            1,
            True,
            Decimal("800"),
            Decimal("200"),
            Decimal("1000"),
            "CNY",
            None,
            None,
            None,
            None,
            captured,
        )
        flights = [
            replace(
                base,
                flight_signature=f"candidate-{index}",
                flight_codes=(f"MU00{index}",),
                total_price_cny=Decimal(990 + index * 10),
            )
            for index in range(1, 7)
        ]
        references = {
            flight.flight_signature: PreferredPriceReference(
                first_total_price_cny=flight.total_price_cny + Decimal("100"),
                first_captured_at=datetime(2026, 9, 1, 8),
                previous_total_price_cny=flight.total_price_cny + Decimal("20"),
                previous_captured_at=datetime(2026, 9, 2, 10, 30),
            )
            for flight in flights
        }
        result = LegResult(
            route(),
            LegStatus.SUCCESS,
            captured,
            flights=flights,
            flight_price_references=references,
            completed_response=True,
        )
        value = RunReport("run-five", captured, captured, RunStatus.SUCCESS, [result], set())
        message = build_message(value, MAIL, sender="sender@example.com", recipients=["to@example.com"])
        plain = message.get_body(preferencelist=("plain",)).get_content()
        html_body = message.get_body(preferencelist=("html",)).get_content()
        self.assertIn("最低价候选（前5条）", plain)
        self.assertIn("[5]", plain)
        self.assertNotIn("[6]", plain)
        self.assertGreaterEqual(plain.count("较首次：下降 ¥100"), 5)
        self.assertGreaterEqual(plain.count("较上次：下降 ¥20"), 5)
        self.assertIn("#5", html_body)
        self.assertNotIn("MU006", html_body)
        self.assertGreaterEqual(html_body.count("较首次：下降 ¥100"), 5)
        self.assertGreaterEqual(html_body.count("较上次：下降 ¥20"), 5)

    def test_sqlite_and_workbook_roundtrip(self):
        with tempfile.TemporaryDirectory() as directory:
            store = SQLiteStore(Path(directory) / "monitor.sqlite3")
            store.initialize()
            value = report()
            store.save_report(value)
            history = store.history(since=datetime(2026, 8, 30))
            self.assertEqual(len(history), 1)
            output = generate_workbook(value, history, Path(directory) / "outputs")
            workbook = load_workbook(output, read_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["本次汇总", "航程1", "24小时历史"])
                headers = [cell.value for cell in next(workbook["航程1"].iter_rows())]
                self.assertIn("行程类型", headers)
                self.assertIn("中转机场", headers)
                self.assertIn("中转等待(分钟)", headers)
            finally:
                workbook.close()

    def test_sqlite_migrates_existing_snapshot_table_for_connections(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Path(directory) / "monitor.sqlite3"
            with closing(sqlite3.connect(database)) as connection:
                connection.execute(
                    "CREATE TABLE flight_snapshots (run_id TEXT, leg_id TEXT, flight_signature TEXT)"
                )
                connection.commit()
            store = SQLiteStore(database)
            store.initialize()
            with closing(store.connect()) as connection:
                columns = {row[1] for row in connection.execute("PRAGMA table_info(flight_snapshots)")}
            self.assertIn("connection_airports_json", columns)
            self.assertIn("layover_minutes", columns)
            self.assertIn("return_itinerary_json", columns)
            self.assertIn("seat_availability_json", columns)
            self.assertIn("outbound_seat_availability_json", columns)
